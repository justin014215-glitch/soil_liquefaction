import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import numpy as np
import os

class LiquefactionExcelReport:
    def __init__(self):
        """初始化Excel報表產生器"""
        self.data = None
        self.workbook = None
        self.worksheet = None

    def load_data_from_dict(self, data_dict):
        """從字典載入資料"""
        self.data = pd.DataFrame(data_dict)
        
        # 檢查必要欄位
        required_fields = ['土層深度', '統一土壤分類']
        missing_fields = [field for field in required_fields if field not in self.data.columns]
        if missing_fields:
            print(f"警告：資料中沒有找到以下欄位：{missing_fields}")
            
        return self.data
    
    def load_data_from_dataframe(self, df):
        """從DataFrame載入資料並過濾到20m深度"""
        # 先過濾到20m深度
        filtered_df = self.filter_data_to_20m(df.copy())
        self.data = filtered_df
        
        # 檢查必要欄位
        required_fields = ['土層深度', '統一土壤分類']
        missing_fields = [field for field in required_fields if field not in self.data.columns]
        if missing_fields:
            print(f"警告：資料中沒有找到以下欄位：{missing_fields}")
            
        return self.data
    
    def filter_data_to_20m(self, df):
        """過濾資料到20m深度，取到第一個超過20m的值並設為20m"""
        if df.empty:
            return df
        
        # 按鑽孔編號分組處理
        filtered_groups = []
        
        for hole_id in df['鑽孔編號'].unique():
            hole_data = df[df['鑽孔編號'] == hole_id].copy()
            
            # 檢查土層深度欄位
            if '土層深度' not in hole_data.columns:
                print(f"警告：鑽孔 {hole_id} 沒有土層深度欄位")
                filtered_groups.append(hole_data)
                continue
            
            # 找到第一個超過20m的索引
            first_over_20_idx = None
            for i, (_, row) in enumerate(hole_data.iterrows()):
                depth = row['土層深度']
                if pd.notna(depth) and float(depth) > 20:
                    first_over_20_idx = i
                    break
            
            if first_over_20_idx is not None:
                # 保留到第一個超過20m的層（包含）
                filtered_hole_data = hole_data.iloc[:first_over_20_idx + 1].copy()
                
                # 將第一個超過20m的土層深度設為20
                last_idx = filtered_hole_data.index[-1]
                filtered_hole_data.loc[last_idx, '土層深度'] = 20.0
                
                print(f"鑽孔 {hole_id}: 保留 {len(filtered_hole_data)} 層，最後一層深度調整為20m")
                filtered_groups.append(filtered_hole_data)
            else:
                # 沒有超過20m的層，全部保留
                print(f"鑽孔 {hole_id}: 全部保留 {len(hole_data)} 層（無超過20m的土層）")
                filtered_groups.append(hole_data)
        
        if filtered_groups:
            result_df = pd.concat(filtered_groups, ignore_index=True)
            print(f"過濾結果：從 {len(df)} 筆資料縮減為 {len(result_df)} 筆資料")
            return result_df
        else:
            return df
    
    def create_excel_report(self, filename='液化潛能分析報表.xlsx'):
        """建立Excel報表 (.xlsx格式)"""
        if self.data is None:
            print("錯誤：請先載入資料")
            return
            
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "液化潛能分析"
        
        # 設定字型樣式
        header_font = Font(name='標楷體', size=10, bold=True)
        data_font = Font(name='標楷體', size=9)
        
        # 設定對齊方式
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 設定邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === 新增標題區塊 ===
        # 取得第一筆資料作為標題資訊
        first_row = self.data.iloc[0] if len(self.data) > 0 else {}

        # 新增標題字型
        title_font = Font(name='標楷體', size=9, bold=True)
        left_alignment = Alignment(horizontal='left', vertical='center')
        title_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        # 第1行
        self.worksheet['A1'] = "計畫名稱"
        self.worksheet['B1'] = first_row.get('計畫名稱', '國震中心SPT液化評估方法測試案例')
        self.worksheet['F1'] = "鑽孔編號"
        self.worksheet['G1'] = first_row.get('鑽孔編號', '')

        # 第2行
        self.worksheet['A2'] = "程式開發"
        self.worksheet['B2'] = "中興工程顧問有限公司"
        self.worksheet['D2'] = "分析單位"
        self.worksheet['E2'] = "中興工程顧問大地工程部"
        self.worksheet['H2'] = "分析方式"
        self.worksheet['I2'] = "HBF (2022)"

        # 第3行 - 座標資訊
        self.worksheet['A3'] = "座標(TWD97)"
        self.worksheet['B3'] = "x:"
        self.worksheet['C3'] = first_row.get('TWD97_X', '')
        self.worksheet['D3'] = "y:"
        self.worksheet['E3'] = first_row.get('TWD97_Y', '')
        self.worksheet['H3'] = "分析地下水位"
        self.worksheet['I3'] = f"{first_row.get('GWT_CSR', 1)}m"

        # 設定標題區塊樣式
        for row in range(1, 4):
            for col in range(1, 23):
                cell = self.worksheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.fill = title_fill
                if cell.value and isinstance(cell.value, str) and any(x in str(cell.value) for x in ['計畫名稱', '程式開發', '座標', '鑽孔編號', '分析方式', '分析單位', '分析地下水位']):
                    cell.font = title_font
                else:
                    cell.font = data_font
                cell.alignment = left_alignment

        # 設定背景色
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # 取得實際的PGA和Mw值（從第一筆資料）
        pga_mid = first_row.get('A_value_MidEq', 0.086)
        mw_mid = first_row.get('Mw_MidEq', 7.1)
        pga_design = first_row.get('A_value_Design', 0.24)
        mw_design = first_row.get('Mw_Design', 7.3)
        pga_max = first_row.get('A_value_MaxEq', 0.32)
        mw_max = first_row.get('Mw_MaxEq', 7.5)
        
        # 建立表頭（四層結構）
        headers = [
            # 第一層 - 主要分類
            ['深度(m)', '柱狀圖', 'SPT-N值', '計算深度(m)', '基本資料', '', '', '', '', 
             f'中小地震PGA={pga_mid}; Mw={mw_mid}', '', '', '', 
             f'設計地震PGA={pga_design}; Mw={mw_design}', '', '', '', 
             f'最大地震PGA={pga_max}; Mw={mw_max}', '', '', '', '安全係數(FS)'],
            # 第二層 - 細項分類
            ['', '', '', '', 'N值', 'γt', 'σv', 'σv\'', 'FC', 'CRR', 'CSR', 'FS', 'LPI',
             'CRR', 'CSR', 'FS', 'LPI', 'CRR', 'CSR', 'FS', 'LPI', ''],
            # 第三層 - 單位
            ['', '', '', '', '', '(t/m³)', '(t/m²)', '(t/m²)', '(%)', '', '', '', '',
             '', '', '', '', '', '', '', '', ''],
            # 第四層 - 空白行（用於格式對齊）
            ['', '', '', '', '', '', '', '', '', '', '', '', '',
             '', '', '', '', '', '', '', '', '']
        ]
        
        # 寫入表頭
        start_row = 5  # 從第5行開始寫表頭
        for row_idx, header_row in enumerate(headers, start_row):
            for col_idx, header in enumerate(header_row, 1):
                cell = self.worksheet.cell(row=row_idx, column=col_idx, value=header)
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.fill = header_fill
        
        # 合併儲存格 - 調整後的範圍
        # 第一行合併
        self.worksheet.merge_cells('A5:A8')  # 深度(m)
        self.worksheet.merge_cells('B5:B8')  # 柱狀圖
        self.worksheet.merge_cells('C5:C8')  # SPT-N值
        self.worksheet.merge_cells('D5:D8')  # 計算深度(m)
        self.worksheet.merge_cells('E5:I5')  # 基本資料
        self.worksheet.merge_cells('J5:M5')  # 中小地震
        self.worksheet.merge_cells('N5:Q5')  # 設計地震
        self.worksheet.merge_cells('R5:U5')  # 最大地震
        self.worksheet.merge_cells('V5:V8')  # 安全係數(FS)
        
        # 寫入資料
        current_row = 9  # 從第9行開始寫資料
        for idx, row in self.data.iterrows():
            # 深度範圍 (第1欄) - 留空，之後放圖表
            self.worksheet.cell(row=current_row, column=1, value='')
            
            # 柱狀圖 (第2欄) - 留空，之後放圖表
            self.worksheet.cell(row=current_row, column=2, value='')
            
            # SPT-N值 (第3欄) - 使用原始N_value
            self.worksheet.cell(row=current_row, column=3, value='')
            
            # 計算深度 (第4欄) - 使用土層深度
            self.worksheet.cell(row=current_row, column=4, value=row.get('土層深度', ''))
            
            # 基本資料
            self.worksheet.cell(row=current_row, column=5, value=row.get('N', ''))  # N值
            self.worksheet.cell(row=current_row, column=6, value=row.get('統體單位重(t/m3)', ''))  # γt
            self.worksheet.cell(row=current_row, column=7, value=row.get('累計sigmav', ''))  # σv
            self.worksheet.cell(row=current_row, column=8, value=row.get('sigma_v_CSR', ''))  # σv'
            self.worksheet.cell(row=current_row, column=9, value=row.get('FC', ''))  # FC
            
            # 中小地震 (MidEq)
            self.worksheet.cell(row=current_row, column=10, value=row.get('CRR_MidEq', ''))
            self.worksheet.cell(row=current_row, column=11, value=row.get('CSR_MidEq', ''))
            self.worksheet.cell(row=current_row, column=12, value=row.get('FS_MidEq', ''))
            self.worksheet.cell(row=current_row, column=13, value=row.get('LPI_MidEq', ''))
            
            # 設計地震 (Design)
            self.worksheet.cell(row=current_row, column=14, value=row.get('CRR_Design', ''))
            self.worksheet.cell(row=current_row, column=15, value=row.get('CSR_Design', ''))
            self.worksheet.cell(row=current_row, column=16, value=row.get('FS_Design', ''))
            self.worksheet.cell(row=current_row, column=17, value=row.get('LPI_Design', ''))
            
            # 最大地震 (MaxEq)
            self.worksheet.cell(row=current_row, column=18, value=row.get('CRR_MaxEq', ''))
            self.worksheet.cell(row=current_row, column=19, value=row.get('CSR_MaxEq', ''))
            self.worksheet.cell(row=current_row, column=20, value=row.get('FS_MaxEq', ''))
            self.worksheet.cell(row=current_row, column=21, value=row.get('LPI_MaxEq', ''))
            
            # 安全係數圖表欄位 (第22欄) - 留空，之後放圖表
            self.worksheet.cell(row=current_row, column=22, value='')
            
            current_row += 1
        
        # 在資料結束後計算並添加LPI總和
        # 計算各情境的LPI總和
        lpi_columns = {
            'M': 'LPI_MidEq',    # 中小地震LPI
            'Q': 'LPI_Design',   # 設計地震LPI  
            'U': 'LPI_MaxEq'     # 最大地震LPI
        }
        
        lpi_sums = {}
        for col_letter, data_col in lpi_columns.items():
            lpi_values = []
            for _, row in self.data.iterrows():
                lpi_val = row.get(data_col, '')
                if lpi_val != '' and lpi_val != '-' and pd.notna(lpi_val):
                    try:
                        lpi_values.append(float(lpi_val))
                    except (ValueError, TypeError):
                        continue
            
            if lpi_values:
                total_lpi = sum(lpi_values)
                lpi_sums[col_letter] = round(total_lpi, 3)
            else:
                lpi_sums[col_letter] = 0.0
        
        # 添加分隔行
        separator_row = current_row
        for col in range(1, 23):
            cell = self.worksheet.cell(row=separator_row, column=col)
            cell.border = Border(
                top=Side(style='medium'),
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        current_row += 1
        
        # 添加LPI總和行
        sum_row = current_row
        
        # 標籤
        label_cell = self.worksheet.cell(row=sum_row, column=5, value="LPI總和")
        label_cell.font = Font(name='標楷體', size=9, bold=True)
        label_cell.alignment = center_alignment
        label_cell.border = thin_border
        label_cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # 合併標籤欄位
        self.worksheet.merge_cells(f'E{sum_row}:I{sum_row}')
        
        # 填入各情境的LPI總和
        for col_letter, sum_value in lpi_sums.items():
            sum_cell = self.worksheet.cell(row=sum_row, column=self.worksheet[col_letter + '1'].column, value=sum_value)
            sum_cell.font = Font(name='標楷體', size=9, bold=True)
            sum_cell.alignment = center_alignment
            sum_cell.border = thin_border
            sum_cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        
        # 調整欄寬
        column_widths = {
            'A': 12, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12,
            'J': 7, 'K': 7, 'L': 7, 'M': 7, 'N': 7, 'O': 7, 'P': 7, 'Q': 7,
            'R': 7, 'S': 7, 'T': 7, 'U': 7, 'V': 15,
        }
        
        for col_letter, width in column_widths.items():
            self.worksheet.column_dimensions[col_letter].width = width

        # 儲存Excel 2007格式檔案
        self.workbook.save(filename)
        print(f"✅ Excel報表已儲存為 {filename}")


class LiquefactionChartGenerator:
    def __init__(self, n_chart_size=(10, 8), fs_chart_size=(12, 8), soil_chart_size=(4, 10)):
        """初始化圖表生成器"""
        self.n_chart_size = n_chart_size        # N值圖表大小
        self.fs_chart_size = fs_chart_size      # FS圖表大小
        self.soil_chart_size = soil_chart_size  # 土壤柱狀圖大小
        self.setup_chinese_font()
        
    def setup_chinese_font(self):
        """設定中文字型"""
        try:
            # 設定matplotlib參數
            plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei', 'SimHei', 'Arial Unicode MS', 'DejaVu Sans']
            plt.rcParams['axes.unicode_minus'] = False
            plt.rcParams['figure.facecolor'] = 'white'
            plt.rcParams['axes.facecolor'] = 'white'
            print("✅ 字型設定完成")
        except Exception as e:
            print(f"⚠️ 字型設定失敗：{e}")
    
    def generate_depth_n_chart(self, data_df, hole_id, output_dir):
        """生成計算深度 vs N值的折線圖"""
        try:
            # 檢查輸出目錄
            if not output_dir or not os.path.exists(output_dir):
                print(f"  ❌ 輸出目錄無效：{output_dir}")
                return None
            
            print(f"    開始生成深度-N值圖表...")
            print(f"    資料筆數：{len(data_df)}")
            
            # 提取資料 - 改善資料提取邏輯
            depths = []
            n_values = []
            
            # 檢查可用的欄位名稱
            depth_columns = ['土層深度', '分析點深度', '計算深度', '深度']
            n_columns = ['N', 'N值', 'SPT_N', 'N_value', 'spt_n']
            
            # 尋找深度欄位
            depth_col = None
            for col in depth_columns:
                if col in data_df.columns:
                    depth_col = col
                    break
            
            # 尋找N值欄位
            n_col = None
            for col in n_columns:
                if col in data_df.columns:
                    n_col = col
                    break
            
            if depth_col is None:
                print(f"    ❌ 找不到深度欄位，可用欄位：{list(data_df.columns)}")
                return None
                
            if n_col is None:
                print(f"    ❌ 找不到N值欄位，可用欄位：{list(data_df.columns)}")
                return None
            
            print(f"    使用深度欄位：{depth_col}")
            print(f"    使用N值欄位：{n_col}")
            
            # 提取有效資料，並過濾深度超過20m的資料點
            for idx, row in data_df.iterrows():
                depth_val = row.get(depth_col)
                n_val = row.get(n_col)
                
                # 處理深度值
                if pd.notna(depth_val) and depth_val != "" and depth_val != "-":
                    try:
                        depth_num = float(depth_val)
                        # 【新增】只保留0-20m的資料
                        if 0 <= depth_num <= 20:
                            # 處理N值
                            if pd.notna(n_val) and n_val != "" and n_val != "-":
                                try:
                                    # 處理 ">50" 這類情況
                                    if isinstance(n_val, str) and n_val.startswith('>'):
                                        n_num = float(n_val[1:])
                                    else:
                                        n_num = float(n_val)
                                    
                                    if n_num >= 0:  # 確保N值為正值
                                        depths.append(depth_num)
                                        n_values.append(n_num)
                                        
                                except (ValueError, TypeError):
                                    continue
                    except (ValueError, TypeError):
                        continue
            
            print(f"    有效資料點數：{len(depths)}")
            
            if not depths or not n_values:
                print(f"    ⚠️ 鑽孔 {hole_id} 沒有有效的深度-N值資料")
                return None
            
            # 顯示資料範圍
            print(f"    深度範圍：{min(depths):.2f} ~ {max(depths):.2f} m")
            print(f"    N值範圍：{min(n_values):.2f} ~ {max(n_values):.2f}")
            
            # 建立圖表
            fig, ax = plt.subplots(figsize=self.n_chart_size)
            
            # 繪製折線圖
            ax.plot(n_values, depths, 'bo-', linewidth=2, markersize=6, 
                   label='SPT-N值', color='blue')
            
            # 設定圖表屬性
            ax.set_xlabel('SPT-N值', fontsize=12, fontweight='bold')
            ax.set_ylabel('深度 (m)', fontsize=12, fontweight='bold')
            ax.set_title(f'鑽孔 {hole_id} - SPT-N值隨深度變化圖', fontsize=14, fontweight='bold')
            
            # 反轉Y軸（深度向下增加）
            ax.invert_yaxis()
            
            # 設定格線
            ax.grid(True, alpha=0.3)
            
            # 【修改】固定軸範圍
            ax.set_xlim(0, max(50, max(n_values) + 5))
            ax.set_ylim(20, 0)  # 固定Y軸範圍為0-20m
            
            # 添加圖例
            ax.legend(fontsize=10)
            
            # 調整布局
            plt.tight_layout()
            
            # 儲存圖表
            filename = f"{hole_id}_深度N值關係圖.jpg"
            filepath = os.path.join(output_dir, filename)
            print(f"    儲存路徑：{filepath}")
            
            plt.savefig(filepath, dpi=300, bbox_inches='tight', 
                       facecolor='white', edgecolor='none')
            plt.close()
            
            print(f"  ✅ 已生成：{filename}")
            return filepath
            
        except Exception as e:
            print(f"  ❌ 生成深度-N值圖表失敗：{e}")
            import traceback
            print(f"     詳細錯誤：{traceback.format_exc()}")
            plt.close()
            return None

    def generate_depth_fs_chart(self, data_df, hole_id, output_dir):
        """生成計算深度 vs 三種情境FS的折線圖"""
        try:
            # 檢查輸出目錄
            if not output_dir or not os.path.exists(output_dir):
                print(f"  ❌ 輸出目錄無效：{output_dir}")
                return None
            
            print(f"    開始生成深度-FS圖表...")
            print(f"    資料筆數：{len(data_df)}")
            
            # 提取資料並過濾深度超過20m的資料點
            depths = []
            fs_mid = []
            fs_design = []
            fs_max = []
            
            # 檢查可用的欄位名稱
            depth_columns = ['土層深度', '分析點深度', '計算深度', '深度']
            
            # 尋找深度欄位
            depth_col = None
            for col in depth_columns:
                if col in data_df.columns:
                    depth_col = col
                    break
            
            if depth_col is None:
                print(f"    ❌ 找不到深度欄位，可用欄位：{list(data_df.columns)}")
                return None
            
            print(f"    使用深度欄位：{depth_col}")
            
            # 檢查FS欄位是否存在
            fs_columns = ['FS_MidEq', 'FS_Design', 'FS_MaxEq']
            missing_fs_cols = [col for col in fs_columns if col not in data_df.columns]
            if missing_fs_cols:
                print(f"    ❌ 找不到FS欄位：{missing_fs_cols}")
                print(f"    可用欄位：{list(data_df.columns)}")
                return None
            
            # 提取有效資料，並過濾深度超過20m的資料點
            for idx, row in data_df.iterrows():
                depth_val = row.get(depth_col)
                fs_m = row.get('FS_MidEq')
                fs_d = row.get('FS_Design')
                fs_x = row.get('FS_MaxEq')
                
                # 處理深度值
                if pd.notna(depth_val) and depth_val != "" and depth_val != "-":
                    try:
                        depth_num = float(depth_val)
                        # 【新增】只保留0-20m的資料
                        if 0 <= depth_num <= 20:
                            depths.append(depth_num)
                            
                            # 處理FS值，將"-"或無效值轉為None
                            def process_fs(fs_val):
                                if fs_val == "-" or pd.isna(fs_val) or fs_val == "":
                                    return None
                                try:
                                    fs_num = float(fs_val)
                                    return fs_num if fs_num >= 0 else None
                                except (ValueError, TypeError):
                                    return None
                            
                            fs_mid.append(process_fs(fs_m))
                            fs_design.append(process_fs(fs_d))
                            fs_max.append(process_fs(fs_x))
                            
                    except (ValueError, TypeError):
                        continue
            
            print(f"    有效資料點數：{len(depths)}")
            
            if not depths:
                print(f"    ⚠️ 鑽孔 {hole_id} 沒有有效的深度-FS資料")
                return None
            
            # 統計有效FS值
            valid_fs_mid = [fs for fs in fs_mid if fs is not None]
            valid_fs_design = [fs for fs in fs_design if fs is not None]
            valid_fs_max = [fs for fs in fs_max if fs is not None]
            
            print(f"    中小地震有效FS數：{len(valid_fs_mid)}")
            print(f"    設計地震有效FS數：{len(valid_fs_design)}")
            print(f"    最大地震有效FS數：{len(valid_fs_max)}")
            
            # 建立圖表
            fig, ax = plt.subplots(figsize=self.fs_chart_size)
            
            # 繪製三條折線（只繪製有效數據）
            scenarios = [
                (fs_mid, '中小地震', 'blue', 'o'),
                (fs_design, '設計地震', 'green', 's'),
                (fs_max, '最大地震', 'red', '^')
            ]
            
            plotted_any = False
            for fs_values, label, color, marker in scenarios:
                # 只繪製有效的數據點
                valid_depths = []
                valid_fs = []
                
                for d, fs in zip(depths, fs_values):
                    if fs is not None and fs >= 0:
                        valid_depths.append(d)
                        valid_fs.append(fs)
                
                if valid_depths:
                    ax.plot(valid_fs, valid_depths, marker + '-', 
                           linewidth=2, markersize=6, color=color, 
                           label=label, alpha=0.8)
                    plotted_any = True
                    print(f"    {label}: 繪製 {len(valid_depths)} 個點")
            
            if not plotted_any:
                print(f"    ⚠️ 沒有任何有效的FS資料可以繪製")
                plt.close()
                return None
            
            # 添加安全係數=1.0的參考線
            ax.axvline(x=1.0, color='black', linestyle='--', alpha=0.7, 
                      label='FS=1.0 (液化臨界值)')
            
            # 設定圖表屬性
            ax.set_xlabel('安全係數 (FS)', fontsize=12, fontweight='bold')
            ax.set_ylabel('深度 (m)', fontsize=12, fontweight='bold')
            ax.set_title(f'鑽孔 {hole_id} - 安全係數隨深度變化圖', fontsize=14, fontweight='bold')
            
            # 設定格線
            ax.grid(True, alpha=0.3)
            
            # 反轉Y軸（深度向下增加）
            ax.invert_yaxis()
            
            # 【修改】設定軸範圍
            ax.set_xlim(0, 3.0)
            ax.set_xticks([0, 1.0, 2.0, 3.0])
            ax.set_xticklabels(['0', '1.0', '2.0', '3.0'])
            ax.set_ylim(20, 0)  # 固定Y軸範圍為0-20m
            
            print(f"    深度範圍：{min(depths):.2f} ~ {max(depths):.2f} m")
            
            # 添加圖例
            ax.legend(fontsize=10, loc='best')
            
            # 調整布局
            plt.tight_layout()
            
            # 儲存圖表
            filename = f"{hole_id}_深度安全係數關係圖.jpg"
            filepath = os.path.join(output_dir, filename)
            print(f"    儲存路徑：{filepath}")
            
            plt.savefig(filepath, dpi=300, bbox_inches='tight', 
                       facecolor='white', edgecolor='none')
            plt.close()
            
            print(f"  ✅ 已生成：{filename}")
            return filepath
            
        except Exception as e:
            print(f"  ❌ 生成深度-FS圖表失敗：{e}")
            import traceback
            print(f"     詳細錯誤：{traceback.format_exc()}")
            plt.close()
            return None
    
    def generate_soil_column_chart(self, data_df, hole_id, output_dir):
        """生成土壤柱狀圖"""
        try:
            # 檢查輸出目錄
            if not output_dir or not os.path.exists(output_dir):
                print(f"  ❌ 輸出目錄無效：{output_dir}")
                return None
            
            print(f"    開始生成土壤柱狀圖...")
            
            # 土壤分類顏色對應表（根據您提供的圖片）
            soil_colors = {
                'GP': '#87CEEB',    # 淺藍色 - 級配良好礫石
                'GW': '#87CEEB',    # 淺藍色 - 級配良好礫石  
                'GM': '#87CEEB',    # 淺藍色 - 粉土質礫石
                'SM': '#FFA500',    # 橘色 - 粉土質砂土
                'SP': '#FFA500',    # 橘色 - 級配不良砂土
                'SW': '#FFA500',    # 橘色 - 級配良好砂土
                'ML': '#C0C0C0',    # 灰色 - 低塑性粉土
                'CL-ML': '#C0C0C0', # 灰色 - 粉土質黏土
                'CL': '#000000',    # 黑色 - 低塑性黏土
                'SC': '#000000',    # 黑色 - 黏土質砂土
                'CH': '#000000',    # 黑色 - 高塑性黏土
                'MH': '#000000',    # 黑色 - 高塑性粉土
            }
            
            # 檢查必要欄位
            depth_columns = ['土層深度', '分析點深度', '計算深度', '深度']
            soil_columns = ['統一土壤分類', '土壤分類', 'USCS']
            
            depth_col = None
            for col in depth_columns:
                if col in data_df.columns:
                    depth_col = col
                    break
            
            soil_col = None
            for col in soil_columns:
                if col in data_df.columns:
                    soil_col = col
                    break
            
            if depth_col is None or soil_col is None:
                print(f"    ❌ 找不到必要欄位")
                return None
            
            # 提取資料，包含第一個超過20m的土層
            soil_layers = []
            found_over_20 = False
            
            for idx, row in data_df.iterrows():
                depth_val = row.get(depth_col)
                soil_val = row.get(soil_col)
                
                if pd.notna(depth_val) and pd.notna(soil_val):
                    try:
                        depth_num = float(depth_val)
                        
                        # 如果深度在20m以內，直接加入
                        if depth_num <= 20:
                            soil_layers.append({
                                'depth': depth_num,
                                'soil_type': str(soil_val).strip(),
                                'upper_depth': row.get('上限深度(公尺)', 0),
                                'lower_depth': row.get('下限深度(公尺)', depth_num)
                            })
                        # 如果是第一個超過20m的土層，也要加入但深度限制在20m
                        elif depth_num > 20 and not found_over_20:
                            soil_layers.append({
                                'depth': 20.0,  # 限制深度在20m
                                'soil_type': str(soil_val).strip(),
                                'upper_depth': row.get('上限深度(公尺)', 0),
                                'lower_depth': 20.0  # 限制在20m
                            })
                            found_over_20 = True
                            print(f"    包含第一個超過20m的土層：{soil_val} (深度限制在20m)")
                            break  # 找到第一個超過20m的就停止
                            
                    except (ValueError, TypeError):
                        continue
            
            if not soil_layers:
                print(f"    ⚠️ 沒有有效的土壤層資料")
                return None
            
            print(f"    土層數量：{len(soil_layers)}")
            
            # 獲取地下水位深度
            gwt_depth = 2.0  # 預設值
            if 'GWT_CSR' in data_df.columns:
                try:
                    gwt_val = data_df['GWT_CSR'].iloc[0]
                    if pd.notna(gwt_val):
                        gwt_depth = float(gwt_val)
                except:
                    pass
            elif 'water_depth(m)' in data_df.columns:
                try:
                    gwt_val = data_df['water_depth(m)'].iloc[0]
                    if pd.notna(gwt_val):
                        gwt_depth = float(gwt_val)
                except:
                    pass
            
            print(f"    地下水位深度：{gwt_depth}m")
            
            # 建立圖表 - 使用可調整的大小
            fig, ax = plt.subplots(figsize=self.soil_chart_size)
            
            # 設定圖表範圍
            ax.set_xlim(0.5, 2.5)
            ax.set_ylim(20, 0)  # Y軸反轉
            
            # 繪製土層
            prev_depth = 0
            for i, layer in enumerate(soil_layers):
                current_depth = layer['depth']
                soil_type = layer['soil_type']
                
                # 確定起始和結束深度
                if i == 0:
                    start_depth = 0
                else:
                    start_depth = prev_depth
                
                end_depth = min(current_depth, 20)
                
                # 獲取土壤顏色
                color = soil_colors.get(soil_type.upper(), '#CCCCCC')
                
                # 繪製土層矩形
                if end_depth > start_depth:
                    rect = plt.Rectangle((0.5, start_depth), 1.5, end_depth - start_depth,
                                       facecolor=color, edgecolor='black', linewidth=0.5)
                    ax.add_patch(rect)
                    print(f"    繪製土層 {i+1}: {soil_type} ({start_depth:.1f}m - {end_depth:.1f}m)")
                
                prev_depth = end_depth
                
                # 如果已經到達20m就停止
                if end_depth >= 20:
                    break
            
            # 繪製地下水位箭頭
            if 0 <= gwt_depth <= 20:
                # 藍色三角形箭頭
                triangle_x = [2.15, 2, 2.15]
                triangle_y = [gwt_depth - 0.2, gwt_depth, gwt_depth + 0.2]
                ax.plot(triangle_x, triangle_y, 'b-', linewidth=2)
                ax.fill(triangle_x, triangle_y, color='blue', alpha=0.7)
                
                # 地下水位標籤
                ax.text(2.2, gwt_depth, f'GWT\n{gwt_depth}m', 
                       fontsize=8, ha='left', va='center', color='blue')
            
            # 設定Y軸刻度和標籤
            y_ticks = list(range(0, 21, 1))
            ax.set_yticks(y_ticks)
            ax.set_yticklabels([str(i) for i in y_ticks])
            
            # 移除X軸刻度
            ax.set_xticks([])
            
            # 設定標籤
            ax.set_ylabel('深度 (m)', fontsize=12, fontweight='bold')
            ax.set_title(f'鑽孔 {hole_id} - 土壤柱狀圖', fontsize=14, fontweight='bold')
            
            # 添加圖例
            legend_elements = []
            unique_soils = set()
            for layer in soil_layers:
                soil_type = layer['soil_type'].upper()
                if soil_type not in unique_soils:
                    unique_soils.add(soil_type)
                    color = soil_colors.get(soil_type, '#CCCCCC')
                    legend_elements.append(plt.Rectangle((0, 0), 1, 1, 
                                                       facecolor=color, 
                                                       edgecolor='black',
                                                       label=soil_type))
            
            # 只顯示前6個圖例項目以避免過度擁擠
            """
            if len(legend_elements) > 6:
                legend_elements = legend_elements[:6]
            """
            if legend_elements:
                ax.legend(handles=legend_elements, loc='center left', 
                         bbox_to_anchor=(1, 0.5), fontsize=8)
            
            # 設定格線
            ax.grid(True, alpha=0.3, axis='y')
            
            # 調整布局
            plt.tight_layout()
            
            # 儲存圖表
            filename = f"{hole_id}_土壤柱狀圖.jpg"
            filepath = os.path.join(output_dir, filename)
            
            plt.savefig(filepath, dpi=300, bbox_inches='tight', 
                       facecolor='white', edgecolor='none')
            plt.close()
            
            print(f"  ✅ 已生成：{filename}")
            return filepath
            
        except Exception as e:
            print(f"  ❌ 生成土壤柱狀圖失敗：{e}")
            import traceback
            print(f"     詳細錯誤：{traceback.format_exc()}")
            plt.close()
            return None

    def generate_charts_for_well(self, data_df, hole_id, output_dir):
        """
        為單個鑽孔生成所有圖表
        
        參數:
        data_df: 包含單個鑽孔資料的DataFrame
        hole_id: 鑽孔編號
        output_dir: 輸出資料夾路徑
        
        返回:
        生成的圖表檔案路徑列表
        """
        generated_files = []
        
        print(f"  正在為鑽孔 {hole_id} 生成圖表...")
        print(f"  資料形狀：{data_df.shape}")
        print(f"  可用欄位：{list(data_df.columns)}")
        
        # 生成深度-N值圖表
        chart1 = self.generate_depth_n_chart(data_df, hole_id, output_dir)
        if chart1:
            generated_files.append(chart1)
        
        # 生成深度-FS圖表
        chart2 = self.generate_depth_fs_chart(data_df, hole_id, output_dir)
        if chart2:
            generated_files.append(chart2)
        
        # 生成土壤柱狀圖
        chart3 = self.generate_soil_column_chart(data_df, hole_id, output_dir)
        if chart3:
            generated_files.append(chart3)
        
        return generated_files
def generate_all_wells_charts(final_df, output_dir,
                              n_chart_size  = (10,8),
                              fs_chart_size = (12,8),
                              soil_chart_size = (5,10)):
    """
    為所有鑽孔生成圖表
    
    參數:
    final_df: 包含所有鑽孔資料的DataFrame
    output_dir: 輸出資料夾路徑
    n_chart_size: N值圖表大小 (寬, 高)
    fs_chart_size: FS圖表大小 (寬, 高)
    返回:
    成功生成的圖表檔案清單
    """
    # 檢查並處理輸出目錄
    if not output_dir or output_dir.strip() == '':
        output_dir = os.getcwd()
        print(f"⚠️ 輸出目錄為空，使用當前工作目錄：{output_dir}")
    
    # 確保輸出目錄存在
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
            print(f"✅ 已創建輸出目錄：{output_dir}")
        except Exception as e:
            print(f"❌ 無法創建輸出目錄：{e}")
            output_dir = os.getcwd()
            print(f"改用當前工作目錄：{output_dir}")
    
    # 建立圖表子資料夾
    charts_dir = os.path.join(output_dir, "圖表")
    try:
        if not os.path.exists(charts_dir):
            os.makedirs(charts_dir)
            print(f"✅ 已創建圖表目錄：{charts_dir}")
    except Exception as e:
        print(f"❌ 無法創建圖表目錄：{e}")
        charts_dir = output_dir  # 如果無法創建子目錄，就使用主目錄
        print(f"圖表將儲存在主目錄：{charts_dir}")
    
    chart_generator = LiquefactionChartGenerator(
                n_chart_size = n_chart_size,
                fs_chart_size = fs_chart_size,
                soil_chart_size = soil_chart_size
    )
    generated_files = []
    well_ids = final_df['鑽孔編號'].unique()
    
    print(f"\n=== 正在生成 {len(well_ids)} 個鑽孔的圖表 ===")
    print(f"圖表輸出目錄：{charts_dir}")
    print(f"總資料筆數：{len(final_df)}")
    print(f"資料欄位：{list(final_df.columns)}")
    
    for i, well_id in enumerate(well_ids, 1):
        try:
            print(f"\n進度 [{i}/{len(well_ids)}] 生成鑽孔 {well_id} 的圖表...")
            
            # 篩選該鑽孔的資料
            well_data = final_df[final_df['鑽孔編號'] == well_id].copy()
            
            if len(well_data) == 0:
                print(f"  警告：鑽孔 {well_id} 沒有資料")
                continue
            
            print(f"  鑽孔 {well_id} 資料筆數：{len(well_data)}")
            
            # 檢查關鍵欄位是否存在
            key_columns = ['土層深度', 'N', 'FS_Design', 'FS_MidEq', 'FS_MaxEq']
            missing_cols = [col for col in key_columns if col not in well_data.columns]
            if missing_cols:
                print(f"  警告：缺少關鍵欄位：{missing_cols}")
                # 嘗試尋找替代欄位
                if '土層深度' not in well_data.columns:
                    alt_depth_cols = ['分析點深度', '計算深度', '深度']
                    for alt_col in alt_depth_cols:
                        if alt_col in well_data.columns:
                            well_data['土層深度'] = well_data[alt_col]
                            print(f"  使用 {alt_col} 作為土層深度")
                            break
            
            # 顯示一些資料統計
            if '土層深度' in well_data.columns:
                depths = well_data['土層深度'].dropna()
                if len(depths) > 0:
                    print(f"  深度範圍：{depths.min():.2f} ~ {depths.max():.2f} m")
            
            if 'N' in well_data.columns:
                n_vals = pd.to_numeric(well_data['N'], errors='coerce').dropna()
                if len(n_vals) > 0:
                    print(f"  N值範圍：{n_vals.min():.2f} ~ {n_vals.max():.2f}")
            
            # 生成圖表
            well_charts = chart_generator.generate_charts_for_well(
                well_data, well_id, charts_dir)
            generated_files.extend(well_charts)
            
            if well_charts:
                print(f"  ✅ 成功生成 {len(well_charts)} 個圖表")
            else:
                print(f"  ⚠️ 未生成任何圖表")
            
        except Exception as e:
            print(f"  ❌ 生成鑽孔 {well_id} 圖表時發生錯誤：{e}")
            import traceback
            print(f"     詳細錯誤：{traceback.format_exc()}")
            continue
    
    print(f"\n✅ 總共成功生成 {len(generated_files)} 個圖表檔案")
    print(f"📁 圖表儲存位置：{charts_dir}")
    
    # 列出生成的檔案
    if generated_files:
        print(f"\n生成的圖表檔案：")
        for file in generated_files:
            print(f"  - {os.path.basename(file)}")
    
    return generated_files

def create_liquefaction_excel(data_dict, filename='液化分析報表.xlsx'):
    """
    簡化版本：直接產生Excel報表
    
    參數:
    data_dict: 包含所有資料的字典
    filename: 輸出檔案名稱
    """
    report = LiquefactionExcelReport()
    report.load_data_from_dict(data_dict)
    report.create_excel_report(filename)
    return filename

def create_liquefaction_excel_from_dataframe(df, filename='液化分析報表.xlsx'):
    """
    從DataFrame產生Excel報表
    
    參數:
    df: 包含單個鑽孔所有資料的DataFrame
    filename: 輸出檔案名稱
    """
    report = LiquefactionExcelReport()
    report.load_data_from_dataframe(df)
    report.create_excel_report(filename)
    return filename

def generate_all_wells_excel_reports(final_df, output_dir):
    """
    為所有鑽孔生成Excel報表
    
    參數:
    final_df: 包含所有鑽孔資料的DataFrame
    output_dir: 輸出資料夾路徑
    
    返回:
    成功生成的報表檔案清單
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    generated_files = []
    well_ids = final_df['鑽孔編號'].unique()
    
    print(f"\n=== 正在生成 {len(well_ids)} 個鑽孔的Excel報表 ===")
    
    for i, well_id in enumerate(well_ids, 1):
        try:
            print(f"進度 [{i}/{len(well_ids)}] 生成鑽孔 {well_id} 的報表...")
            
            # 篩選該鑽孔的資料
            well_data = final_df[final_df['鑽孔編號'] == well_id].copy()
            
            if len(well_data) == 0:
                print(f"  警告：鑽孔 {well_id} 沒有資料")
                continue
            
            # 生成檔案名稱
            filename = f"{well_id}_液化分析報表.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            # 生成報表
            report_file = create_liquefaction_excel_from_dataframe(well_data, filepath)
            generated_files.append(report_file)
            
            print(f"  ✅ 已生成：{filename}")
            
        except Exception as e:
            print(f"  ❌ 生成鑽孔 {well_id} 報表時發生錯誤：{e}")
            continue
    
    print(f"\n✅ 總共成功生成 {len(generated_files)} 個Excel報表")
    print(f"📁 報表儲存位置：{output_dir}")
    
    return generated_files